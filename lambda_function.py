import boto3
import json
import os
import csv
from io import StringIO
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from typing import Dict, List, Any, Tuple

logger = logging.getLogger()
logger.setLevel(logging.INFO)

def clean_column_name(column: str) -> str:
    """Make column names Redshift-friendly"""
    if not column:
        return "unnamed_column"
    cleaned = ''.join(c if c.isalnum() else '_' for c in str(column).lower())
    # Handle columns that start with numbers
    if cleaned[0].isdigit():
        cleaned = 'col_' + cleaned
    return cleaned

def detect_data_type(values: List[Any]) -> str:
    """
    Detect the most appropriate Redshift data type based on column values
    """
    has_number = False
    has_decimal = False
    has_date = False
    max_length = 0
    
    for val in values:
        if val is None or val == '':
            continue
            
        # Check if it's a date
        if isinstance(val, datetime):
            has_date = True
        
        # Check if it's a number
        if isinstance(val, (int, float)):
            has_number = True
            if isinstance(val, float):
                has_decimal = True
        
        # Check string length
        if isinstance(val, str):
            max_length = max(max_length, len(val))
            
            # Try to parse as date
            try:
                datetime.strptime(val, '%Y-%m-%d')
                has_date = True
            except:
                pass
    
    # Determine data type
    if has_date:
        return 'TIMESTAMP'
    elif has_decimal:
        return 'DECIMAL(18,2)'
    elif has_number:
        return 'BIGINT'
    else:
        # For text, use VARCHAR with appropriate length
        length = max(max_length * 2, 100)  # Double length for safety
        length = min(length, 65535)  # Redshift VARCHAR limit
        return f'VARCHAR({length})'

def process_worksheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, str]]:
    """
    Process a single worksheet and return headers, data, and column types
    """
    data = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Get headers from first row
    headers = []
    header_cells = {}
    for col in range(1, max_col + 1):
        cell_value = sheet.cell(row=1, column=col).value
        header = clean_column_name(cell_value if cell_value else f"col_{col}")
        # Handle duplicate column names
        base_header = header
        counter = 1
        while header in headers:
            header = f"{base_header}_{counter}"
            counter += 1
        headers.append(header)
        header_cells[col] = header
    
    # Extract column data for type detection
    column_values = {header: [] for header in headers}
    
    # Process data rows
    for row in range(2, max_row + 1):
        row_data = {}
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            header = header_cells[col]
            value = cell.value
            
            # Handle dates specifically
            if cell.is_date and value:
                value = value.isoformat() if hasattr(value, 'isoformat') else value
                
            row_data[header] = value
            column_values[header].append(value)
        data.append(row_data)
    
    # Detect column types
    column_types = {header: detect_data_type(values) for header, values in column_values.items()}
    
    return headers, data, column_types

def create_table_if_not_exists(redshift_data: boto3.client, 
                             table_name: str, 
                             headers: List[str], 
                             column_types: Dict[str, str]) -> None:
    """Create table in Redshift if it doesn't exist"""
    columns = [f"{header} {column_types[header]}" for header in headers]
    create_table_sql = f"""
    CREATE TABLE IF NOT EXISTS {table_name} (
        {','.join(columns)}
    );
    """
    
    response = redshift_data.execute_statement(
        Database=os.environ['REDSHIFT_DATABASE'],
        WorkgroupName=os.environ['REDSHIFT_WORKGROUP'],
        Sql=create_table_sql
    )
    logger.info(f"Created table {table_name} with response: {response}")

def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    try:
        # Get S3 file details
        bucket = event['Records'][0]['s3']['bucket']['name']
        key = event['Records'][0]['s3']['object']['key']
        
        # Get the Excel file from S3
        s3 = boto3.client('s3')
        response = s3.get_object(Bucket=bucket, Key=key)
        
        # Load workbook
        wb = openpyxl.load_workbook(response['Body'], data_only=True)
        
        # Process each worksheet
        redshift_data = boto3.client('redshift-data')
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row < 2:  # Skip empty sheets
                continue
                
            # Process worksheet
            headers, data, column_types = process_worksheet(sheet)
            
            # Generate table name
            base_table_name = clean_column_name(os.path.splitext(os.path.basename(key))[0])
            table_name = f"{base_table_name}_{clean_column_name(sheet_name)}"
            
            # Create table
            create_table_if_not_exists(redshift_data, table_name, headers, column_types)
            
            # Convert data to CSV
            csv_buffer = StringIO()
            writer = csv.DictWriter(csv_buffer, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
            
            # Upload CSV to S3
            csv_key = f"temp/{table_name}.csv"
            s3.put_object(
                Bucket=bucket,
                Key=csv_key,
                Body=csv_buffer.getvalue().encode('utf-8')
            )
            
            # COPY command to load data
            copy_command = f"""
            COPY {table_name}
            FROM 's3://{bucket}/{csv_key}'
            IAM_ROLE '{os.environ['REDSHIFT_ROLE_ARN']}'
            CSV HEADER;
            """
            
            # Execute COPY command
            redshift_data.execute_statement(
                Database=os.environ['REDSHIFT_DATABASE'],
                WorkgroupName=os.environ['REDSHIFT_WORKGROUP'],
                Sql=copy_command
            )
            
            # Clean up temporary CSV
            s3.delete_object(Bucket=bucket, Key=csv_key)
            
            logger.info(f"Processed worksheet {sheet_name} into table {table_name}")
        
        return {
            'statusCode': 200,
            'body': json.dumps(f'Successfully processed all worksheets from {key}')
        }
        
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps(f'Error processing file: {str(e)}')
        }